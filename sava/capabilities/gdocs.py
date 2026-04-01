"""Google Workspace capability — read/write docs, sheets, PDFs; manage comments; suggest edits."""

from googleapiclient.discovery import build

from ..auth import get_credentials
from ..browser import run_playwright_action


def _docs_service():
    return build("docs", "v1", credentials=get_credentials())


def _drive_service():
    return build("drive", "v3", credentials=get_credentials())


def _sheets_service():
    return build("sheets", "v4", credentials=get_credentials())


def _download_file(drive, file_id):
    """Download a file from Drive into an in-memory BytesIO."""
    import io
    from googleapiclient.http import MediaIoBaseDownload

    fh = io.BytesIO()
    request = drive.files().get_media(fileId=file_id)
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    fh.seek(0)
    return fh


def read_doc(doc_id: str) -> str:
    """Read the full text content of a file on Google Drive.
    Supports: Google Docs, Google Sheets, Word (.docx), Excel (.xlsx), and PDF."""
    drive = _drive_service()

    meta = drive.files().get(
        fileId=doc_id, fields="name,mimeType", supportsAllDrives=True
    ).execute()
    mime = meta.get("mimeType", "")
    name = meta.get("name", "(untitled)")

    # --- Native Google Doc ---
    if "google-apps.document" in mime:
        docs = _docs_service()
        doc = docs.documents().get(
            documentId=doc_id,
            suggestionsViewMode="SUGGESTIONS_INLINE",
        ).execute()

        parts = []
        for element in doc.get("body", {}).get("content", []):
            paragraph = element.get("paragraph")
            if not paragraph:
                continue
            for pe in paragraph.get("elements", []):
                text_run = pe.get("textRun")
                if text_run:
                    parts.append(text_run["content"])

        return f"# {name}\n\n{''.join(parts)}"

    # --- Native Google Sheet ---
    if "google-apps.spreadsheet" in mime:
        return _read_google_sheet(doc_id, name)

    # --- Word (.docx) ---
    if "wordprocessingml" in mime:
        import zipfile
        from lxml import etree

        fh = _download_file(drive, doc_id)
        ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
        with zipfile.ZipFile(fh) as z:
            root = etree.fromstring(z.read("word/document.xml"))

        paragraphs = []
        for p in root.findall(".//w:p", ns):
            runs = p.findall(".//w:t", ns)
            text = "".join(t.text or "" for t in runs)
            if text:
                paragraphs.append(text)

        return f"# {name}\n\n" + "\n".join(paragraphs)

    # --- Excel (.xlsx) ---
    if "spreadsheetml" in mime:
        import openpyxl

        fh = _download_file(drive, doc_id)
        wb = openpyxl.load_workbook(fh, read_only=True, data_only=True)
        return _format_workbook(wb, name)

    # --- PDF ---
    if "pdf" in mime:
        import pymupdf

        fh = _download_file(drive, doc_id)
        doc = pymupdf.open(stream=fh.read(), filetype="pdf")
        pages = []
        for i, page in enumerate(doc):
            text = page.get_text()
            if text.strip():
                pages.append(f"--- Page {i + 1} ---\n{text}")
        doc.close()

        return f"# {name}\n\n" + "\n".join(pages) if pages else f"# {name}\n\n(No text found in PDF)"

    return f"# {name}\n\n(Unsupported format: {mime})"


def _read_google_sheet(spreadsheet_id: str, name: str) -> str:
    """Read all sheets from a Google Sheet via the Sheets API."""
    sheets = _sheets_service()
    meta = sheets.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
    sheet_names = [s["properties"]["title"] for s in meta.get("sheets", [])]

    parts = [f"# {name}\n"]
    for sheet_name in sheet_names:
        result = sheets.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range=sheet_name,
        ).execute()
        rows = result.get("values", [])
        if not rows:
            parts.append(f"\n## {sheet_name}\n(empty)")
            continue

        parts.append(f"\n## {sheet_name}\n")
        # Format as a markdown table
        header = rows[0]
        parts.append("| " + " | ".join(str(c) for c in header) + " |")
        parts.append("| " + " | ".join("---" for _ in header) + " |")
        for row in rows[1:]:
            # Pad row to match header length
            padded = row + [""] * (len(header) - len(row))
            parts.append("| " + " | ".join(str(c) for c in padded) + " |")

    return "\n".join(parts)


def _format_workbook(wb, name: str) -> str:
    """Format an openpyxl workbook as readable text."""
    parts = [f"# {name}\n"]
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            parts.append(f"\n## {sheet_name}\n(empty)")
            continue

        parts.append(f"\n## {sheet_name}\n")
        header = [str(c) if c is not None else "" for c in rows[0]]
        parts.append("| " + " | ".join(header) + " |")
        parts.append("| " + " | ".join("---" for _ in header) + " |")
        for row in rows[1:]:
            cells = [str(c) if c is not None else "" for c in row]
            cells += [""] * (len(header) - len(cells))
            parts.append("| " + " | ".join(cells) + " |")

    return "\n".join(parts)


def list_comments(doc_id: str) -> str:
    """List all comments and replies on a Google Doc."""
    drive = _drive_service()
    lines = []
    page_token = None

    while True:
        resp = drive.comments().list(
            fileId=doc_id,
            fields="comments(id,content,quotedFileContent,resolved,author,createdTime,"
                   "replies(id,content,author,createdTime,action)),nextPageToken",
            includeDeleted=False,
            pageSize=100,
            pageToken=page_token,
        ).execute()

        for c in resp.get("comments", []):
            status = "[RESOLVED]" if c.get("resolved") else "[OPEN]"
            author = c.get("author", {}).get("displayName", "?")
            quoted = c.get("quotedFileContent", {}).get("value", "")
            lines.append(f"\nComment {c['id']}  {status}  by {author}  {c.get('createdTime', '')}")
            if quoted:
                lines.append(f'  Quoted: "{quoted}"')
            lines.append(f"  {c['content']}")

            for r in c.get("replies", []):
                r_author = r.get("author", {}).get("displayName", "?")
                action = f" [{r['action']}]" if r.get("action") else ""
                lines.append(f"    -> {r_author}{action}: {r['content']}")

        page_token = resp.get("nextPageToken")
        if not page_token:
            break

    return "\n".join(lines) if lines else "No comments found."


def add_comment(doc_id: str, message: str, quote: str = "") -> str:
    """Post a comment on a Google Doc. Optionally reference a text passage."""
    drive = _drive_service()
    content = f'Re: "{quote}"\n\n{message}' if quote else message

    comment = drive.comments().create(
        fileId=doc_id,
        fields="id",
        body={"content": content},
    ).execute()

    return f"Created comment {comment['id']}"


def reply_to_comment(doc_id: str, comment_id: str, message: str) -> str:
    """Reply to an existing comment."""
    drive = _drive_service()
    reply = drive.replies().create(
        fileId=doc_id,
        commentId=comment_id,
        fields="id",
        body={"content": message},
    ).execute()

    return f"Created reply {reply['id']} on comment {comment_id}"


def resolve_comment(doc_id: str, comment_id: str, message: str = "Resolved.") -> str:
    """Resolve a comment thread."""
    drive = _drive_service()
    drive.replies().create(
        fileId=doc_id,
        commentId=comment_id,
        fields="id",
        body={"content": message, "action": "resolve"},
    ).execute()

    return f"Resolved comment {comment_id}"


def write_sheet(spreadsheet_id: str, range: str, values: list[list[str]]) -> str:
    """Write values to a Google Sheet. Range is in A1 notation e.g. 'Sheet1!A1:C3'."""
    sheets = _sheets_service()
    result = sheets.spreadsheets().values().update(
        spreadsheetId=spreadsheet_id,
        range=range,
        valueInputOption="USER_ENTERED",
        body={"values": values},
    ).execute()
    updated = result.get("updatedCells", 0)
    return f"Updated {updated} cells in {range}"


def anchor_comment(doc_id: str, quote: str, message: str) -> str:
    """Post a comment anchored to specific text using browser automation."""
    return run_playwright_action("anchor_comment", doc_id=doc_id, quote=quote, message=message)


def suggest_edit(doc_id: str, quote: str, replacement: str) -> str:
    """Create a tracked suggestion using Find & Replace in Suggesting mode."""
    return run_playwright_action("suggest_edit", doc_id=doc_id, quote=quote, replacement=replacement)


_MIME_LABELS = {
    "application/vnd.google-apps.document": "Google Doc",
    "application/vnd.google-apps.spreadsheet": "Google Sheet",
    "application/vnd.google-apps.folder": "Folder",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": "Word (.docx)",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "Excel (.xlsx)",
    "application/pdf": "PDF",
}


def list_docs() -> str:
    """List all files accessible to Sava on Google Drive."""
    drive = _drive_service()
    lines = []
    page_token = None

    while True:
        results = drive.files().list(
            pageSize=100,
            fields="files(id,name,mimeType,owners,webViewLink),nextPageToken",
            includeItemsFromAllDrives=True,
            supportsAllDrives=True,
            pageToken=page_token,
        ).execute()

        for f in results.get("files", []):
            owners = ", ".join(o.get("emailAddress", "?") for o in f.get("owners", []))
            mime = f.get("mimeType", "")
            label = _MIME_LABELS.get(mime, mime)
            lines.append(f"{f['name']}  [{label}]")
            lines.append(f"  ID: {f['id']}")
            if owners:
                lines.append(f"  Owner: {owners}")
            lines.append(f"  Link: {f.get('webViewLink', 'N/A')}")
            lines.append("")

        page_token = results.get("nextPageToken")
        if not page_token:
            break

    return "\n".join(lines) if lines else "No files found."
